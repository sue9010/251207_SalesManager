import os
import time
import shutil
import openpyxl

# [변경] 설정 파일 경로 수정 (src 패키지)
from src.config import Config

class ExportManager:
    def __init__(self, datamanager):
        self.datamanager = datamanager

    def export_quote_to_pdf(self, client_info, quote_info, items):
        try:
            import win32com.client
            import pythoncom
        except ImportError:
            return False, "PDF 변환을 위해 pywin32 라이브러리가 필요합니다.\n설치: pip install pywin32", None

        temp_xlsx_path = None
        excel = None
        wb_opened = None
        wb = None

        try:
            client_name = quote_info['client_name']
            mgmt_no = quote_info['mgmt_no']
            
            country = str(client_info.get("국가", "Korea")).upper()
            is_kr = any(x in country for x in ["대한민국", "KR", "한국", "KOREA"])
            
            template_filename = "Quotation_KR.xlsx" if is_kr else "Quotation_EN.xlsx"
            template_path = os.path.join(Config.FORMS_DIR, template_filename)

            if not os.path.exists(template_path):
                return False, f"견적서 양식 파일을 찾을 수 없습니다.\n경로: {template_path}", None

            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            ws["B7"] = client_name
            ws["B8"] = str(client_info.get("담당자", ""))
            ws["B9"] = str(client_info.get("전화번호", ""))
            ws["B10"] = str(client_info.get("주소", ""))

            ws["G10"] = quote_info['date']
            ws["G11"] = mgmt_no

            start_row = 14
            for i, item in enumerate(items):
                if i >= 10: break
                r_idx = start_row + i
                
                ws[f"B{r_idx}"] = item.get('item', '')
                ws[f"C{r_idx}"] = item.get('model', '')
                ws[f"D{r_idx}"] = item.get('desc', '')
                ws[f"E{r_idx}"] = item.get('qty', 0)
                ws[f"F{r_idx}"] = item.get('price', 0)
                ws[f"G{r_idx}"] = item.get('amount', 0)

            ws["B28"] = quote_info.get('note', '')

            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            
            if is_kr:
                pdf_filename = f"견적서_{client_name}_{mgmt_no}.pdf"
            else:
                pdf_filename = f"Quotation_{client_name}_{mgmt_no}.pdf"
            
            success, msg, server_path = self._convert_and_save_pdf(wb, pdf_filename, server_folder_name="견적서")
            
            if success and server_path:
                self.datamanager.update_order_fields(mgmt_no, {"견적서경로": server_path})
                
            return success, msg, server_path

        except Exception as e:
            return False, f"오류 발생: {str(e)}", None
        
        finally:
            if wb:
                try: wb.close()
                except: pass

    def export_order_request_to_pdf(self, client_info, order_info, items):
        try:
            import win32com.client
            import pythoncom
        except ImportError:
            return False, "PDF 변환을 위해 pywin32 라이브러리가 필요합니다.", None

        temp_xlsx_path = None
        excel = None
        wb_opened = None
        wb = None

        try:
            template_path = os.path.join(Config.FORMS_DIR, "order_request_form.xlsx")
            if not os.path.exists(template_path):
                return False, f"출고요청서 양식 파일을 찾을 수 없습니다.\n경로: {template_path}", None

            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            ws["C1"] = order_info.get('type', '')
            ws["C2"] = order_info.get('mgmt_no', '')
            ws["C3"] = order_info.get('date', '')
            
            ws["C7"] = str(client_info.get("주소", ""))
            
            ws["E1"] = order_info.get('client_name', '')
            ws["E3"] = str(client_info.get("국가", ""))
            ws["E4"] = str(client_info.get("전화번호", ""))
            ws["E5"] = str(client_info.get("운송계정", ""))
            ws["E6"] = str(client_info.get("운송방법", ""))

            start_row = 10
            for i, item in enumerate(items):
                if i >= 10: break 
                r_idx = start_row + i
                
                ws[f"C{r_idx}"] = item.get('item', '')
                ws[f"D{r_idx}"] = item.get('model', '')
                ws[f"E{r_idx}"] = item.get('desc', '')
                ws[f"F{r_idx}"] = item.get('qty', 0)

            ws["B21"] = order_info.get('req_note', '')
            ws["B24"] = str(client_info.get("특이사항", ""))

            client_safe = "".join([c for c in order_info['client_name'] if c.isalnum() or c in (' ', '_')]).strip()
            pdf_filename = f"출고요청서_{client_safe}_{order_info['mgmt_no']}.pdf"

            success, msg, server_path = self._convert_and_save_pdf(wb, pdf_filename, server_folder_name="출고요청서")
            
            if success and server_path:
                self.datamanager.update_order_fields(order_info['mgmt_no'], {"출고요청서경로": server_path})
                
            return success, msg, server_path

        except Exception as e:
            return False, f"오류 발생: {str(e)}", None
            
        finally:
            if wb:
                try: wb.close()
                except: pass

    def export_pi_to_pdf(self, client_info, order_info, items):
        try:
            import win32com.client
            import pythoncom
        except ImportError:
            return False, "PDF 변환을 위해 pywin32 라이브러리가 필요합니다.", None

        temp_xlsx_path = None
        excel = None
        wb_opened = None
        wb = None

        try:
            template_path = os.path.join(Config.FORMS_DIR, "PI.xlsx")
            
            if not os.path.exists(template_path):
                return False, f"PI 양식 파일을 찾을 수 없습니다.\n경로: {template_path}", None

            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            ws["B5"] = order_info.get('client_name', '')
            ws["B6"] = str(client_info.get("담당자", ""))
            ws["B7"] = str(client_info.get("전화번호", ""))
            ws["B8"] = str(client_info.get("주소", ""))
            ws["G9"] = str(client_info.get("국가", ""))

            ws["G5"] = order_info.get('date', '')
            ws["G6"] = order_info.get('mgmt_no', '')
            ws["G7"] = order_info.get('po_no', '')

            start_row = 12
            for i, item in enumerate(items):
                if i >= 10: break
                r_idx = start_row + i
                
                ws[f"B{r_idx}"] = item.get('item', '')
                ws[f"C{r_idx}"] = item.get('model', '')
                ws[f"D{r_idx}"] = item.get('desc', '')
                ws[f"E{r_idx}"] = item.get('qty', 0)
                ws[f"F{r_idx}"] = item.get('price', 0)
                ws[f"G{r_idx}"] = item.get('amount', 0)

            client_safe = "".join([c for c in order_info['client_name'] if c.isalnum() or c in (' ', '_', '-')]).strip()
            pdf_filename = f"PI_{client_safe}_{order_info['mgmt_no']}.pdf"
            
            success, msg, server_path = self._convert_and_save_pdf(wb, pdf_filename, server_folder_name="Proforma Invoice")
            
            if success and server_path:
                self.datamanager.update_order_fields(order_info['mgmt_no'], {"PI경로": server_path})
                
            return success, msg, server_path

        except Exception as e:
            return False, f"오류 발생: {str(e)}", None
            
        finally:
            if wb:
                try: wb.close()
                except: pass

    def export_ci_to_pdf(self, client_info, order_info, items):
        try:
            import win32com.client
            import pythoncom
        except ImportError:
            return False, "PDF 변환을 위해 pywin32 라이브러리가 필요합니다.", None

        temp_xlsx_path = None
        excel = None
        wb_opened = None
        wb = None

        try:
            template_path = os.path.join(Config.FORMS_DIR, "CI.xlsx")
            
            if not os.path.exists(template_path):
                return False, f"CI 양식 파일을 찾을 수 없습니다.\n경로: {template_path}", None

            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            def safe_write(cell_addr, val):
                try:
                    for merged_range in ws.merged_cells.ranges:
                        if cell_addr in merged_range:
                            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            top_left_cell.value = val
                            return
                    ws[cell_addr] = val
                except: pass

            safe_write("A9", order_info.get('client_name', ''))
            safe_write("A10", str(client_info.get("주소", "")))
            safe_write("A12", str(client_info.get("전화번호", "")))
            safe_write("A13", str(client_info.get("담당자", "")))
            safe_write("E13", str(client_info.get("국가", "")))

            safe_write("E9", order_info.get('mgmt_no', ''))
            safe_write("E11", order_info.get('date', ''))

            po_list = []
            if order_info.get('po_no'):
                po_list.append(str(order_info.get('po_no')))
            for item in items:
                p = str(item.get('po_no', '')).strip()
                if p and p != "nan" and p not in po_list:
                    po_list.append(p)
            final_po_str = ", ".join(list(dict.fromkeys(po_list))) 
            safe_write("E10", final_po_str)

            serial_list = []
            
            start_row = 16
            for i, item in enumerate(items):
                if i >= 10: break
                r_idx = start_row + i
                
                safe_write(f"A{r_idx}", item.get('model', ''))
                safe_write(f"B{r_idx}", item.get('desc', ''))
                safe_write(f"C{r_idx}", item.get('qty', 0))
                safe_write(f"D{r_idx}", item.get('currency', ''))
                safe_write(f"E{r_idx}", item.get('price', 0))
                safe_write(f"H{r_idx}", item.get('currency', ''))
                safe_write(f"I{r_idx}", item.get('amount', 0))
                
                sn = str(item.get('serial', '')).strip()
                if sn and sn != "-" and sn != "nan":
                    serial_list.append(sn)

            if serial_list:
                safe_write("A32", ", ".join(serial_list))

            client_safe = "".join([c for c in order_info['client_name'] if c.isalnum() or c in (' ', '_', '-')]).strip()

            # [변경] 송장번호가 있으면 파일명에 포함
            invoice_no = order_info.get('invoice_no', '')
            if invoice_no:
                pdf_filename = f"CI_{client_safe}_{order_info['mgmt_no']}_{invoice_no}.pdf"
            else:
                pdf_filename = f"CI_{client_safe}_{order_info['mgmt_no']}.pdf"

            return self._convert_and_save_pdf(wb, pdf_filename, server_folder_name="Commercial Invoice")

        except Exception as e:
            return False, f"오류 발생: {str(e)}", None
            
        finally:
            if wb:
                try: wb.close()
                except: pass

    def export_pl_to_pdf(self, client_info, order_info, items):
        try:
            import win32com.client
            import pythoncom
        except ImportError:
            return False, "PDF 변환을 위해 pywin32 라이브러리가 필요합니다.", None

        temp_xlsx_path = None
        excel = None
        wb_opened = None
        wb = None

        try:
            template_path = os.path.join(Config.FORMS_DIR, "PL.xlsx")
            
            if not os.path.exists(template_path):
                return False, f"PL 양식 파일을 찾을 수 없습니다.\n경로: {template_path}", None

            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            def safe_write(cell_addr, val):
                try:
                    for merged_range in ws.merged_cells.ranges:
                        if cell_addr in merged_range:
                            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            top_left_cell.value = val
                            return
                    ws[cell_addr] = val
                except: pass

            safe_write("A9", order_info.get('client_name', ''))
            safe_write("A10", str(client_info.get("주소", "")))
            safe_write("A12", str(client_info.get("전화번호", "")))
            safe_write("A13", str(client_info.get("담당자", "")))
            
            safe_write("F9", order_info.get('mgmt_no', ''))
            safe_write("F11", order_info.get('date', ''))
            safe_write("F13", str(client_info.get("국가", "")))

            po_list = []
            if order_info.get('po_no'): po_list.append(str(order_info.get('po_no')))
            for item in items:
                p = str(item.get('po_no', '')).strip()
                if p and p != "nan" and p not in po_list:
                    po_list.append(p)
            final_po_str = ", ".join(list(dict.fromkeys(po_list))) 
            safe_write("F10", final_po_str)

            serial_list = []
            total_net_weight = 0.0
            total_gross_weight = 0.0
            carton_count = 0

            start_row = 16
            for i, item in enumerate(items):
                if i >= 10: break
                r_idx = start_row + i
                
                safe_write(f"A{r_idx}", item.get('c_no', ''))
                safe_write(f"B{r_idx}", item.get('desc', ''))
                safe_write(f"C{r_idx}", item.get('qty', 0))
                safe_write(f"D{r_idx}", item.get('unit', 'SET'))
                
                nw = item.get('net_weight', 0)
                gw = item.get('gross_weight', 0)
                
                try: nw_float = float(str(nw).replace("kg","").strip())
                except: nw_float = 0
                try: gw_float = float(str(gw).replace("kg","").strip())
                except: gw_float = 0
                
                total_net_weight += nw_float
                total_gross_weight += gw_float
                
                if item.get('c_no'): carton_count += 1
                
                safe_write(f"E{r_idx}", nw)
                safe_write(f"F{r_idx}", gw)
                
                safe_write(f"G{r_idx}", item.get('size_l', ''))
                safe_write(f"H{r_idx}", item.get('size_w', ''))
                safe_write(f"I{r_idx}", item.get('size_h', ''))
                
                sn = str(item.get('serial', '')).strip()
                if sn and sn != "-" and sn != "nan":
                    serial_list.append(sn)

            if carton_count == 0 and items: carton_count = len(items)
            safe_write("A26", f"{carton_count} Cartons")
            safe_write("E26", f"{total_net_weight:.2f} kgs")
            safe_write("F26", f"{total_gross_weight:.2f} kgs")

            safe_write("A29", order_info.get('notes', ''))
            if serial_list:
                safe_write("A36", ", ".join(serial_list))

            client_safe = "".join([c for c in order_info['client_name'] if c.isalnum() or c in (' ', '_', '-')]).strip()

            # [변경] 송장번호가 있으면 파일명에 포함
            invoice_no = order_info.get('invoice_no', '')
            if invoice_no:
                pdf_filename = f"PL_{client_safe}_{order_info['mgmt_no']}_{invoice_no}.pdf"
            else:
                pdf_filename = f"PL_{client_safe}_{order_info['mgmt_no']}.pdf"

            return self._convert_and_save_pdf(wb, pdf_filename, server_folder_name="Packing List")

        except Exception as e:
            return False, f"오류 발생: {str(e)}", None
            
        finally:
            if wb:
                try: wb.close()
                except: pass

    def _convert_and_save_pdf(self, wb, pdf_filename, server_folder_name=None):
        """
        공통 PDF 변환 및 저장 로직
        1. 바탕화면에 임시 xlsx 저장
        2. Excel로 PDF 변환 (바탕화면)
        3. server_folder_name이 있으면 해당 서버 폴더로 복사
        4. 임시 파일 삭제 및 정리
        """
        temp_xlsx_path = None
        excel = None
        wb_opened = None
        
        try:
            import win32com.client
            import pythoncom
        except ImportError:
            return False, "PDF 변환을 위해 pywin32 라이브러리가 필요합니다.", None

        try:
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            
            # 임시 엑셀 파일명 (유니크하게)
            import uuid
            temp_xlsx_name = f"temp_{uuid.uuid4().hex[:8]}.xlsx"
            temp_xlsx_path = os.path.join(desktop, temp_xlsx_name)
            
            wb.save(temp_xlsx_path)
            
            desktop_pdf_path = os.path.join(desktop, pdf_filename)
            
            # PDF 변환
            pythoncom.CoInitialize()
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False

            abs_xlsx_path = os.path.abspath(temp_xlsx_path)
            wb_opened = excel.Workbooks.Open(abs_xlsx_path)
            
            try:
                wb_opened.ExportAsFixedFormat(0, desktop_pdf_path)
            except Exception as e:
                return False, f"PDF 변환 실패: {e}", None
            
            wb_opened.Close(SaveChanges=False)
            wb_opened = None
            
            msg = f"바탕화면에 저장되었습니다.\n{desktop_pdf_path}"
            
            # 서버 복사
            if server_folder_name:
                # [변경] DataManager에서 설정된 attachment_root 사용
                server_dir = os.path.join(self.datamanager.attachment_root, server_folder_name)
                server_pdf_path = os.path.join(server_dir, pdf_filename)
                
                copied_to_server = False
                
                # 서버 폴더 생성 시도
                if not os.path.exists(server_dir):
                    try: os.makedirs(server_dir)
                    except: pass
                
                if os.path.exists(server_dir):
                    try:
                        shutil.copy2(desktop_pdf_path, server_pdf_path)
                        copied_to_server = True
                    except Exception as e:
                        print(f"서버 복사 실패: {e}")
                
                if copied_to_server:
                    msg += f"\n\n서버에도 저장되었습니다.\n{server_pdf_path}"
                    return True, msg, server_pdf_path
            
            return True, msg, None

        except Exception as e:
            return False, f"PDF 처리 중 오류 발생: {str(e)}", None
            
        finally:
            if wb_opened:
                try: wb_opened.Close(False)
                except: pass
            if excel:
                try: excel.Quit(); del excel
                except: pass
            if temp_xlsx_path and os.path.exists(temp_xlsx_path):
                try: os.remove(temp_xlsx_path)
                except: pass
