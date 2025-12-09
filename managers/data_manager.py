import getpass
import json
import os
import shutil
from datetime import datetime

import pandas as pd
import openpyxl

# [변경] 설정 파일 경로 수정 (src 패키지)
from src.config import Config


class DataManager:
    def __init__(self):
        self.df_clients = pd.DataFrame(columns=Config.CLIENT_COLUMNS)
        self.df_data = pd.DataFrame(columns=Config.DATA_COLUMNS)
        self.df_payment = pd.DataFrame(columns=Config.PAYMENT_COLUMNS)
        self.df_delivery = pd.DataFrame(columns=Config.DELIVERY_COLUMNS)
        self.df_log = pd.DataFrame(columns=Config.LOG_COLUMNS)
        self.df_memo = pd.DataFrame(columns=Config.MEMO_COLUMNS)
        self.df_memo_log = pd.DataFrame(columns=Config.MEMO_LOG_COLUMNS)
        
        self.current_excel_path = Config.DEFAULT_EXCEL_PATH
        self.attachment_root = Config.DEFAULT_ATTACHMENT_ROOT
        self.production_request_path = Config.DEFAULT_PRODUCTION_REQUEST_PATH
        self.purchase_data_path = Config.DEFAULT_PURCHASE_DATA_PATH
        self.order_request_dir = Config.DEFAULT_ORDER_REQUEST_DIR 
        
        self.current_theme = "Dark"
        self.is_dev_mode = False
        self.last_file_timestamp = 0.0
        
        self.load_config()

    def load_config(self):
        if os.path.exists(Config.CONFIG_FILENAME):
            try:
                with open(Config.CONFIG_FILENAME, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.current_excel_path = data.get("excel_path", Config.DEFAULT_EXCEL_PATH)
                    self.current_theme = data.get("theme", "Dark")
                    self.attachment_root = data.get("attachment_root", Config.DEFAULT_ATTACHMENT_ROOT)
                    self.production_request_path = data.get("production_request_path", Config.DEFAULT_PRODUCTION_REQUEST_PATH)
                    self.purchase_data_path = data.get("purchase_data_path", Config.DEFAULT_PURCHASE_DATA_PATH)
                    self.order_request_dir = data.get("order_request_dir", Config.DEFAULT_ORDER_REQUEST_DIR)
            except: pass

    def save_config(self, new_path=None, new_theme=None, new_attachment_dir=None, new_prod_path=None, new_order_req_dir=None, new_purchase_path=None):
        if new_path: self.current_excel_path = new_path
        if new_theme: self.current_theme = new_theme
        if new_attachment_dir: self.attachment_root = new_attachment_dir
        if new_prod_path: self.production_request_path = new_prod_path
        if new_purchase_path: self.purchase_data_path = new_purchase_path
        if new_order_req_dir: self.order_request_dir = new_order_req_dir
        
        data = {
            "excel_path": self.current_excel_path,
            "theme": self.current_theme,
            "attachment_root": self.attachment_root,
            "production_request_path": self.production_request_path,
            "purchase_data_path": self.purchase_data_path,
            "order_request_dir": self.order_request_dir
        }
        try:
            with open(Config.CONFIG_FILENAME, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"설정 저장 실패: {e}")

    # ==========================================================================
    # Helper Methods (Task 1-A)
    # ==========================================================================
    def _get_columns_for_key(self, key):
        if key == "clients": return Config.CLIENT_COLUMNS
        elif key == "data": return Config.DATA_COLUMNS
        elif key == "payment": return Config.PAYMENT_COLUMNS
        elif key == "delivery": return Config.DELIVERY_COLUMNS
        elif key == "log": return Config.LOG_COLUMNS
        elif key == "memo": return Config.MEMO_COLUMNS
        elif key == "memo_log": return Config.MEMO_LOG_COLUMNS
        return []

    def _generate_sequential_id(self, df, id_col, prefix):
        today = datetime.now().strftime("%y%m%d")
        prefix_date = f"{prefix}-{today}-"
        
        max_seq = 0
        if not df.empty and id_col in df.columns:
            def extract_seq(val):
                s = str(val)
                if s.startswith(prefix_date):
                    try: return int(s.split("-")[-1])
                    except: return 0
                return 0
            
            seqs = df[id_col].apply(extract_seq)
            if not seqs.empty: max_seq = seqs.max()
        
        next_seq = max_seq + 1
        return f"{prefix_date}{next_seq:03d}"

    def _read_all_sheets(self) -> dict[str, pd.DataFrame]:
        dfs = {}
        if not os.path.exists(self.current_excel_path):
            keys = ["clients", "data", "payment", "delivery", "log", "memo", "memo_log"]
            for key in keys:
                dfs[key] = pd.DataFrame(columns=self._get_columns_for_key(key))
            return dfs

        try:
            with open(self.current_excel_path, "rb") as f:
                with pd.ExcelFile(f, engine="openpyxl") as xls:
                    sheet_names = xls.sheet_names
                    
                    sheet_map = {
                        Config.SHEET_CLIENTS: "clients",
                        Config.SHEET_DATA: "data",
                        Config.SHEET_PAYMENT: "payment",
                        Config.SHEET_DELIVERY: "delivery",
                        Config.SHEET_LOG: "log",
                        Config.SHEET_MEMO: "memo",
                        Config.SHEET_MEMO_LOG: "memo_log"
                    }
                    
                    for sheet_name, key in sheet_map.items():
                        if sheet_name in sheet_names:
                            df = pd.read_excel(xls, sheet_name)
                            if key in ["clients", "data"]:
                                df.columns = df.columns.astype(str).str.strip()
                            dfs[key] = df
                        else:
                            dfs[key] = pd.DataFrame(columns=self._get_columns_for_key(key))
                            
            return dfs
        except Exception as e:
            print(f"Error reading sheets: {e}")
            return {}

    def _normalize_all(self, dfs: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
        keys = ["clients", "data", "payment", "delivery", "log", "memo", "memo_log"]
        for key in keys:
            if key not in dfs:
                dfs[key] = pd.DataFrame(columns=self._get_columns_for_key(key))

        if "data" in dfs:
            for col in Config.DATA_COLUMNS:
                if col not in dfs["data"].columns: dfs["data"][col] = "-"
            dfs["data"] = dfs["data"].fillna("-")
            
            num_cols = ["수량", "단가", "환율", "세율(%)", "공급가액", "세액", "합계금액", "기수금액", "미수금액"]
            for col in num_cols:
                if col in dfs["data"].columns:
                    dfs["data"][col] = pd.to_numeric(dfs["data"][col], errors='coerce').fillna(0)

            date_cols = ["견적일", "수주일", "출고예정일", "출고일", "선적일", "입금완료일", "세금계산서발행일"]
            for col in date_cols:
                if col in dfs["data"].columns:
                    dfs["data"][col] = pd.to_datetime(dfs["data"][col], errors='coerce', format='mixed').dt.strftime("%Y-%m-%d")
                    dfs["data"][col] = dfs["data"][col].fillna("-")

        if "clients" in dfs:
            dfs["clients"] = dfs["clients"].fillna("-")

        if "delivery" in dfs:
            if "출고번호" not in dfs["delivery"].columns:
                dfs["delivery"]["출고번호"] = "-"
            for col in Config.DELIVERY_COLUMNS:
                 if col not in dfs["delivery"].columns: dfs["delivery"][col] = "-"
            dfs["delivery"] = dfs["delivery"].fillna("-")

        return dfs

    def _write_all_sheets(self, dfs: dict[str, pd.DataFrame]) -> None:
        with pd.ExcelWriter(self.current_excel_path, engine="openpyxl") as writer:
            sheet_map = {
                "clients": Config.SHEET_CLIENTS,
                "data": Config.SHEET_DATA,
                "payment": Config.SHEET_PAYMENT,
                "delivery": Config.SHEET_DELIVERY,
                "log": Config.SHEET_LOG,
                "memo": Config.SHEET_MEMO,
                "memo_log": Config.SHEET_MEMO_LOG
            }
            for key, sheet_name in sheet_map.items():
                if key in dfs:
                    dfs[key].to_excel(writer, sheet_name=sheet_name, index=False)

    def load_data(self):
        try:
            dfs = self._read_all_sheets()
            dfs = self._normalize_all(dfs)
            
            self.df_clients = dfs["clients"]
            self.df_data = dfs["data"]
            self.df_payment = dfs["payment"]
            self.df_delivery = dfs["delivery"]
            self.df_log = dfs["log"]
            self.df_memo = dfs["memo"]
            self.df_memo_log = dfs["memo_log"]
            
            if os.path.exists(self.current_excel_path):
                self.last_file_timestamp = os.path.getmtime(self.current_excel_path)
                
            return True, "데이터 로드 완료"
        except Exception as e:
            return False, f"오류 발생: {e}"

    def check_for_external_changes(self):
        if not os.path.exists(self.current_excel_path): return False
        try:
            current_mtime = os.path.getmtime(self.current_excel_path)
            if current_mtime > self.last_file_timestamp: return True
        except OSError: pass
        return False

    def _execute_transaction(self, update_logic_func):
        if not os.path.exists(self.current_excel_path):
            return False, "엑셀 파일이 존재하지 않습니다."

        try:
            dfs = self._read_all_sheets()
            dfs = self._normalize_all(dfs)
            
            success, msg = update_logic_func(dfs)
            if not success: return False, msg

            self._write_all_sheets(dfs)
            self.load_data()
            return True, "저장되었습니다."

        except PermissionError:
            return False, "엑셀 파일이 열려있습니다. 파일을 닫고 다시 시도해주세요."
        except Exception as e:
            return False, f"트랜잭션 오류: {e}"

    def recalc_payment_status(self, dfs, mgmt_no):
        pay_df = dfs["payment"]
        if pay_df.empty:
            total_paid = 0
            last_pay_date = "-"
        else:
            target_pays = pay_df[pay_df["관리번호"].astype(str) == str(mgmt_no)]
            paid_series = target_pays["입금액"].astype(str).str.replace(",", "").replace("nan", "0")
            total_paid = pd.to_numeric(paid_series, errors='coerce').sum()
            
            if not target_pays.empty:
                target_pays = target_pays.sort_values(by="일시", ascending=False)
                last_pay_date = target_pays.iloc[0]["일시"].split(" ")[0]
            else:
                last_pay_date = "-"

        data_df = dfs["data"]
        mask = data_df["관리번호"] == mgmt_no
        indices = data_df[mask].index
        
        if len(indices) == 0: return

        amt_series = data_df.loc[mask, "합계금액"].astype(str).str.replace(",", "").replace("nan", "0")
        total_contract_amt = pd.to_numeric(amt_series, errors='coerce').sum()
        
        remaining_to_allocate = total_paid
        
        for idx in indices:
            val_str = str(data_df.at[idx, "합계금액"]).replace(",", "")
            try: row_total = float(val_str)
            except: row_total = 0
            
            if remaining_to_allocate >= row_total:
                allocated = row_total
            else:
                allocated = remaining_to_allocate
                if allocated < 0: allocated = 0
            
            data_df.at[idx, "기수금액"] = allocated
            data_df.at[idx, "미수금액"] = row_total - allocated
            
            remaining_to_allocate -= allocated
            
            try: row_unpaid = float(data_df.at[idx, "미수금액"])
            except: row_unpaid = 0
            
            current_status = str(data_df.at[idx, "Status"])
            
            if row_unpaid < 1:
                if "납품" in current_status or "완료" in current_status:
                    new_status = "완료"
                else:
                    new_status = "납품대기/입금완료"
                data_df.at[idx, "입금완료일"] = last_pay_date
            else:
                if current_status == "완료": new_status = "납품완료/입금완료"
                elif current_status == "납품대기/입금완료": new_status = current_status
                else: new_status = current_status
            
            data_df.at[idx, "Status"] = new_status

    def _create_log_entry(self, action, details):
        try: user = getpass.getuser()
        except: user = "Unknown"
        return {
            "일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "작업자": user,
            "구분": action,
            "상세내용": details
        }

    def save_to_excel(self):
        try:
            with pd.ExcelWriter(self.current_excel_path, engine="openpyxl") as writer:
                self.df_clients.to_excel(writer, sheet_name=Config.SHEET_CLIENTS, index=False)
                self.df_data.to_excel(writer, sheet_name=Config.SHEET_DATA, index=False)
                self.df_payment.to_excel(writer, sheet_name=Config.SHEET_PAYMENT, index=False)
                self.df_delivery.to_excel(writer, sheet_name=Config.SHEET_DELIVERY, index=False)
                self.df_log.to_excel(writer, sheet_name=Config.SHEET_LOG, index=False)
                self.df_memo.to_excel(writer, sheet_name=Config.SHEET_MEMO, index=False)
                self.df_memo_log.to_excel(writer, sheet_name=Config.SHEET_MEMO_LOG, index=False)
            return True, "저장 완료"
        except PermissionError:
            return False, "엑셀 파일이 열려있습니다."
        except Exception as e:
            return False, f"저장 실패: {e}"

    def save_attachment(self, source_path, company_name, file_prefix):
        if not os.path.exists(source_path): return None, "파일 없음"
        try:
            current_year = datetime.now().strftime("%Y")
            year_dir = os.path.join(self.attachment_root, current_year)
            if not os.path.exists(year_dir): os.makedirs(year_dir)

            safe_name = "".join([c for c in str(company_name) if c.isalnum() or c in (' ', '_', '-')]).strip()
            comp_dir = os.path.join(year_dir, safe_name)
            if not os.path.exists(comp_dir): os.makedirs(comp_dir)

            fname = os.path.basename(source_path)
            name, ext = os.path.splitext(fname)
            today = datetime.now().strftime("%Y%m%d")
            new_name = f"{file_prefix}_{today}_{name}{ext}"
            
            dest_path = os.path.join(comp_dir, new_name)
            shutil.copy2(source_path, dest_path)
            return dest_path, None
        except Exception as e:
            return None, str(e)

    def add_log(self, action, details):
        try: user = getpass.getuser()
        except: user = "Unknown"
        new = {"일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "작업자": user, "구분": action, "상세내용": details}
        new_df = pd.DataFrame([new])
        if self.df_log.empty: self.df_log = new_df
        else: self.df_log = pd.concat([self.df_log, new_df], ignore_index=True)
    
    def get_status_by_req_no(self, req_no):
        if self.df_data.empty: return None
        rows = self.df_data[self.df_data["관리번호"] == req_no]
        return rows.iloc[0]["Status"] if not rows.empty else None

    def get_filtered_data(self, status_list=None, keyword=""):
        df = self.df_data
        if df.empty: return df
        if status_list: df = df[df["Status"].isin(status_list)]
        if keyword:
            mask = pd.Series(False, index=df.index)
            for col in Config.SEARCH_TARGET_COLS:
                if col in df.columns:
                    mask |= df[col].astype(str).str.contains(keyword, case=False)
            df = df[mask]
        return df

    def set_dev_mode(self, enabled): self.is_dev_mode = enabled
    
    def create_backup(self):
        if not os.path.exists(self.current_excel_path): return False, "파일 없음"
        try:
            folder = os.path.dirname(self.current_excel_path)
            backup_folder = os.path.join(folder, "backup")
            if not os.path.exists(backup_folder): os.makedirs(backup_folder)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            fname = os.path.basename(self.current_excel_path)
            shutil.copy2(self.current_excel_path, os.path.join(backup_folder, f"{fname}_{timestamp}.bak"))
            return True, "백업 완료"
        except Exception as e: return False, str(e)

    def clean_old_logs(self): return True, "로그 정리 완료"

    def export_to_production_request(self, rows_data):
        prod_path = self.production_request_path
        if not os.path.exists(prod_path):
            return False, f"생산 요청 파일을 찾을 수 없습니다.\n경로: {prod_path}"

        try:
            wb = openpyxl.load_workbook(prod_path)
            if "Data" not in wb.sheetnames:
                return False, "'Data' 시트가 존재하지 않습니다."
            ws = wb["Data"]

            added_count = 0
            updated_count = 0

            for row_data in rows_data:
                client_name = row_data.get("업체명", "")
                client_note = "-"
                if not self.df_clients.empty:
                    c_row = self.df_clients[self.df_clients["업체명"] == client_name]
                    if not c_row.empty:
                        val = c_row.iloc[0].get("특이사항", "-")
                        if str(val) != "nan" and val: client_note = str(val)

                mgmt_no = str(row_data.get("관리번호", ""))
                model_name = str(row_data.get("모델명", ""))
                desc = str(row_data.get("Description", ""))
                order_date = row_data.get("수주일", "-")
                if not order_date or order_date == "nan": order_date = "-"

                mapping_values = [
                    mgmt_no,                    # A
                    client_name,                # B
                    model_name,                 # C
                    desc,                       # D
                    row_data.get("수량", 0),    # E
                    row_data.get("주문요청사항", "-"), # F
                    client_note,                # G
                    order_date,                 # H
                    "-",                        # I
                    "-",                        # J
                    "-",                        # K
                    "-",                        # L
                    "-",                        # M
                    "생산 접수",                # N
                    "-",                        # O
                    "-"                         # P
                ]

                target_row_idx = None
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    curr_mgmt = str(row[0]) if row[0] else ""
                    curr_model = str(row[2]) if row[2] else ""
                    curr_desc = str(row[3]) if row[3] else ""
                    
                    if curr_mgmt == mgmt_no and curr_model == model_name and curr_desc == desc:
                        target_row_idx = i
                        break
                
                if target_row_idx:
                    for col_idx, val in enumerate(mapping_values, start=1):
                        ws.cell(row=target_row_idx, column=col_idx, value=val)
                    updated_count += 1
                else:
                    ws.append(mapping_values)
                    added_count += 1

            wb.save(prod_path)
            wb.close()
            return True, f"신규: {added_count}건, 업데이트: {updated_count}건"

        except PermissionError:
            return False, "생산 요청 파일이 열려있습니다. 파일을 닫고 다시 시도해주세요."
        except Exception as e:
            return False, f"생산 요청 내보내기 실패: {e}"

    def sync_production_dates(self):
        if not os.path.exists(self.production_request_path):
            return

        try:
            wb = openpyxl.load_workbook(self.production_request_path, data_only=True)
            if "Data" not in wb.sheetnames: return
            ws = wb["Data"]
            
            date_map = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                mgmt_no = str(row[0]) if row[0] else None
                delivery_date = row[8]
                
                if mgmt_no and delivery_date:
                    if isinstance(delivery_date, datetime):
                        date_str = delivery_date.strftime("%Y-%m-%d")
                    else:
                        date_str = str(delivery_date).strip()
                        if date_str.lower() == "nan" or date_str == "-" or not date_str:
                            continue
                            
                    date_map[mgmt_no] = date_str
            
            wb.close()
            
            if not self.df_data.empty:
                if '출고예정일' not in self.df_data.columns:
                    self.df_data['출고예정일'] = "-"
                
                for mgmt_no, new_date in date_map.items():
                    mask = self.df_data['관리번호'] == mgmt_no
                    if mask.any():
                        self.df_data.loc[mask, '출고예정일'] = new_date
                        
        except Exception as e:
            print(f"생산 요청일 동기화 실패: {e}")

    def get_production_status_map(self):
        if not os.path.exists(self.production_request_path):
            return {}

        try:
            wb = openpyxl.load_workbook(self.production_request_path, data_only=True, read_only=True)
            if "Data" not in wb.sheetnames:
                return {}
            
            ws = wb["Data"]
            status_map = {}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 14: continue
                
                mgmt_no = str(row[0]).strip() if row[0] else None
                prod_status = str(row[13]).strip() if row[13] else "-"
                
                if mgmt_no:
                    status_map[mgmt_no] = prod_status
            
            wb.close()
            return status_map
        except Exception as e:
            print(f"생산 상태 로드 실패: {e}")
            return {}

    def get_serial_number_map(self):
        if not os.path.exists(self.production_request_path):
            return {}

        try:
            wb = openpyxl.load_workbook(self.production_request_path, data_only=True, read_only=True)
            if "Data" not in wb.sheetnames:
                return {}
            
            ws = wb["Data"]
            serial_map = {}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 11: continue
                
                mgmt_no = str(row[0]).strip() if row[0] else ""
                model = str(row[2]).strip() if row[2] else ""
                desc = str(row[3]).strip() if row[3] else ""
                serial = str(row[10]).strip() if row[10] else "-"
                
                key = (mgmt_no, model, desc)
                if mgmt_no: 
                    serial_map[key] = serial
            
            wb.close()
            return serial_map
        except Exception as e:
            print(f"시리얼 번호 로드 실패: {e}")
            return {}

    def get_client_shipping_method(self, client_name):
        if self.df_clients.empty:
            return ""
            
        row = self.df_clients[self.df_clients["업체명"] == client_name]
        if not row.empty:
            val = row.iloc[0].get("운송방법", "")
            return str(val).strip() if str(val).lower() != "nan" else ""
        return ""

    def get_client_shipping_account(self, client_name):
        if self.df_clients.empty:
            return ""
            
        row = self.df_clients[self.df_clients["업체명"] == client_name]
        if not row.empty:
            val = row.iloc[0].get("운송계정", "")
            return str(val).strip() if str(val).lower() != "nan" else ""
        return ""

    def get_next_quote_id(self):
        return self._generate_sequential_id(self.df_data, "관리번호", "QT")

    def get_next_order_id(self):
        return self._generate_sequential_id(self.df_data, "관리번호", "OD")

    def _add_log_to_dfs(self, dfs, action, details):
        new_log = self._create_log_entry(action, details)
        if "log" in dfs:
            dfs["log"] = pd.concat([dfs["log"], pd.DataFrame([new_log])], ignore_index=True)

    def get_next_delivery_id(self):
        return self._generate_sequential_id(self.df_delivery, "출고번호", "CX")

    def add_client(self, client_data: dict) -> tuple[bool, str]:
        def update(dfs):
            if client_data.get("업체명") in dfs["clients"]["업체명"].values:
                return False, "이미 존재하는 업체명입니다."
            
            new_df = pd.DataFrame([client_data])
            dfs["clients"] = pd.concat([dfs["clients"], new_df], ignore_index=True)
            self._add_log_to_dfs(dfs, "업체 등록", f"업체명: {client_data.get('업체명')}")
            return True, ""
        return self._execute_transaction(update)

    def update_client(self, original_name, client_data: dict) -> tuple[bool, str]:
        def update(dfs):
            mask = dfs["clients"]["업체명"] == original_name
            if not mask.any(): return False, "업체를 찾을 수 없습니다."
            
            idx = dfs["clients"][mask].index[0]
            for k, v in client_data.items():
                dfs["clients"].at[idx, k] = v
                
            self._add_log_to_dfs(dfs, "업체 수정", f"업체명: {original_name}")
            return True, ""
        return self._execute_transaction(update)

    def delete_client(self, client_name) -> tuple[bool, str]:
        def update(dfs):
            mask = dfs["clients"]["업체명"] == client_name
            if not mask.any(): return False, "업체를 찾을 수 없습니다."
            
            dfs["clients"] = dfs["clients"][~mask]
            self._add_log_to_dfs(dfs, "업체 삭제", f"업체명: {client_name}")
            return True, ""
        return self._execute_transaction(update)

    def add_order(self, order_rows: list[dict], mgmt_no: str, client_name: str) -> tuple[bool, str]:
        def update(dfs):
            new_df = pd.DataFrame(order_rows)
            dfs["data"] = pd.concat([dfs["data"], new_df], ignore_index=True)
            self._add_log_to_dfs(dfs, "주문 등록", f"번호 [{mgmt_no}] / 업체 [{client_name}]")
            return True, ""
        return self._execute_transaction(update)

    def update_order(self, mgmt_no: str, order_rows: list[dict], client_name: str, is_copy=False) -> tuple[bool, str]:
        def update(dfs):
            mask = dfs["data"]["관리번호"] == mgmt_no
            existing_rows = dfs["data"][mask]
            
            if not existing_rows.empty:
                first_exist = existing_rows.iloc[0]
                preserve_cols = ["출고예정일", "출고일", "입금완료일", "세금계산서발행일", "계산서번호", "수출신고번호"]
                for row in order_rows:
                    for col in preserve_cols:
                        if col not in row:
                            row[col] = first_exist.get(col, "-")
            
            dfs["data"] = dfs["data"][~mask]
            
            new_df = pd.DataFrame(order_rows)
            dfs["data"] = pd.concat([dfs["data"], new_df], ignore_index=True)
            
            action = "복사 등록" if is_copy else "수정"
            self._add_log_to_dfs(dfs, f"주문 {action}", f"번호 [{mgmt_no}] / 업체 [{client_name}]")
            return True, ""
        return self._execute_transaction(update)

    def delete_order(self, mgmt_no: str) -> tuple[bool, str]:
        def update(dfs):
            mask = dfs["data"]["관리번호"] == mgmt_no
            if not mask.any(): return False, "데이터를 찾을 수 없습니다."
            
            dfs["data"] = dfs["data"][~mask]
            self._add_log_to_dfs(dfs, "삭제", f"주문 삭제: 번호 [{mgmt_no}]")
            return True, ""
        return self._execute_transaction(update)

    def add_payment(self, payment_data: dict) -> tuple[bool, str]:
        def update(dfs):
            new_df = pd.DataFrame([payment_data])
            dfs["payment"] = pd.concat([dfs["payment"], new_df], ignore_index=True)
            
            mgmt_no = payment_data.get("관리번호")
            if mgmt_no:
                self.recalc_payment_status(dfs, mgmt_no)
                
            self._add_log_to_dfs(dfs, "입금 등록", f"관리번호: {mgmt_no}, 금액: {payment_data.get('입금액')}")
            return True, ""
        return self._execute_transaction(update)

    def add_delivery(self, delivery_data: dict) -> tuple[bool, str]:
        def update(dfs):
            new_df = pd.DataFrame([delivery_data])
            dfs["delivery"] = pd.concat([dfs["delivery"], new_df], ignore_index=True)
            self._add_log_to_dfs(dfs, "출고 등록", f"출고번호: {delivery_data.get('출고번호')}")
            return True, ""
        return self._execute_transaction(update)

    def add_quote(self, quote_rows: list[dict], mgmt_no: str, client_name: str) -> tuple[bool, str]:
        def update(dfs):
            new_df = pd.DataFrame(quote_rows)
            dfs["data"] = pd.concat([dfs["data"], new_df], ignore_index=True)
            self._add_log_to_dfs(dfs, "견적 등록", f"번호 [{mgmt_no}] / 업체 [{client_name}]")
            return True, ""
        return self._execute_transaction(update)

    def update_quote(self, mgmt_no: str, quote_rows: list[dict], client_name: str, is_copy=False) -> tuple[bool, str]:
        def update(dfs):
            mask = dfs["data"]["관리번호"] == mgmt_no
            existing_rows = dfs["data"][mask]
            
            if not existing_rows.empty:
                first_exist = existing_rows.iloc[0]
                preserve_cols = ["수주일", "출고예정일", "출고일", "입금완료일", 
                                 "세금계산서발행일", "계산서번호", "수출신고번호", "발주서경로"]
                for row in quote_rows:
                    for col in preserve_cols:
                        if col not in row:
                            row[col] = first_exist.get(col, "-")
            
            dfs["data"] = dfs["data"][~mask]
            
            new_df = pd.DataFrame(quote_rows)
            dfs["data"] = pd.concat([dfs["data"], new_df], ignore_index=True)
            
            action = "복사 등록" if is_copy else "수정"
            self._add_log_to_dfs(dfs, f"견적 {action}", f"번호 [{mgmt_no}] / 업체 [{client_name}]")
            return True, ""
        return self._execute_transaction(update)

    def delete_quote(self, mgmt_no: str) -> tuple[bool, str]:
        def update(dfs):
            mask = dfs["data"]["관리번호"] == mgmt_no
            if not mask.any(): return False, "데이터를 찾을 수 없습니다."
            
            dfs["data"] = dfs["data"][~mask]
            self._add_log_to_dfs(dfs, "삭제", f"견적 삭제: 번호 [{mgmt_no}]")
            return True, ""
        return self._execute_transaction(update)

    def process_delivery(self, delivery_no, delivery_date, invoice_no, shipping_method, waybill_path, update_requests):
        def update(dfs):
            processed_items = []
            new_delivery_records = []
            
            try: current_user = getpass.getuser()
            except: current_user = "Unknown"
            
            for req in update_requests:
                idx = req["idx"]
                if idx not in dfs["data"].index: continue
                
                row_data = dfs["data"].loc[idx]
                db_qty = float(str(row_data["수량"]).replace(",", "") or 0)
                deliver_qty = min(req["deliver_qty"], db_qty)
                
                new_delivery_records.append({
                    "일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "출고번호": delivery_no, "출고일": delivery_date,
                    "관리번호": row_data.get("관리번호", ""), "품목명": row_data.get("품목명", ""),
                    "시리얼번호": req["serial_no"], "출고수량": deliver_qty,
                    "송장번호": invoice_no, "운송방법": shipping_method,
                    "작업자": current_user, "비고": "일괄 납품 처리",
                    "운송장경로": waybill_path
                })

                # 데이터 업데이트 (완전 출고 vs 부분 출고)
                is_full = abs(deliver_qty - db_qty) < 0.000001
                new_status = "완료" if row_data.get("Status") == "납품대기/입금완료" else "납품완료/입금대기"
                
                price = float(str(row_data.get("단가", 0)).replace(",", "") or 0)
                tax_rate = float(str(row_data.get("세율(%)", 0)).replace(",", "") or 0) / 100

                if is_full:
                    dfs["data"].at[idx, "Status"] = new_status
                    dfs["data"].at[idx, "출고일"] = delivery_date
                    dfs["data"].at[idx, "송장번호"] = invoice_no
                    dfs["data"].at[idx, "운송방법"] = shipping_method
                    dfs["data"].at[idx, "미수금액"] = float(str(row_data.get("합계금액", 0)).replace(",", ""))
                else:
                    remain_qty = db_qty - deliver_qty
                    supply = remain_qty * price
                    tax = supply * tax_rate
                    dfs["data"].at[idx, "수량"] = remain_qty
                    dfs["data"].at[idx, "공급가액"] = supply
                    dfs["data"].at[idx, "세액"] = tax
                    dfs["data"].at[idx, "합계금액"] = supply + tax
                    dfs["data"].at[idx, "미수금액"] = supply + tax
                    
                    new_supply = deliver_qty * price
                    new_tax = new_supply * tax_rate
                    new_row = row_data.copy()
                    new_row.update({
                        "수량": deliver_qty, "공급가액": new_supply, "세액": new_tax, "합계금액": new_supply + new_tax,
                        "미수금액": new_supply + new_tax, "Status": new_status, "출고일": delivery_date,
                        "송장번호": invoice_no, "운송방법": shipping_method,
                    })
                    if "운송장경로" in new_row: new_row["운송장경로"] = ""
                    dfs["data"] = pd.concat([dfs["data"], pd.DataFrame([new_row])], ignore_index=True)
                
                processed_items.append(f"{row_data.get('품목명','')} ({deliver_qty}개)")

            if new_delivery_records:
                dfs["delivery"] = pd.concat([dfs["delivery"], pd.DataFrame(new_delivery_records)], ignore_index=True)

            log_msg = f"번호 [{row_data.get('관리번호', '')}...] 납품 처리(출고번호: {delivery_no}) / {', '.join(processed_items)}"
            self._add_log_to_dfs(dfs, "납품 처리", log_msg)
            return True, ""
            
        return self._execute_transaction(update)

    def process_payment(self, mgmt_nos, payment_amount, payment_date, file_paths, confirm_fee_callback=None):
        def update(dfs):
            mask = dfs["data"]["관리번호"].isin(mgmt_nos)
            if not mask.any(): return False, "데이터를 찾을 수 없습니다."

            indices = dfs["data"][mask].index
            
            # 1. 강제 재계산
            for mgmt_no in mgmt_nos:
                self.recalc_payment_status(dfs, mgmt_no)

            # 2. 배치 처리용 집계
            batch_summary = {}
            now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            try: current_user = getpass.getuser()
            except: current_user = "Unknown"

            remaining_payment = payment_amount

            # 3. 미수금 차감 시뮬레이션
            for idx in indices:
                if remaining_payment <= 0: break
                
                mgmt_no = dfs["data"].at[idx, "관리번호"]
                currency = str(dfs["data"].at[idx, "통화"]).upper()
                threshold = 200 if currency != "KRW" else 5000
                
                if mgmt_no not in batch_summary:
                    batch_summary[mgmt_no] = {'deposit': 0, 'fee': 0, 'currency': currency}

                try: unpaid = float(dfs["data"].at[idx, "미수금액"])
                except: unpaid = 0
                
                if unpaid > 0:
                    actual_pay = 0
                    fee_pay = 0
                    
                    if remaining_payment >= unpaid:
                        actual_pay = unpaid
                    else:
                        diff = unpaid - remaining_payment
                        if diff <= threshold:
                            item_name = str(dfs["data"].at[idx, "품목명"])
                            is_fee = False
                            if confirm_fee_callback:
                                is_fee = confirm_fee_callback(item_name, diff, currency)
                            
                            if is_fee:
                                actual_pay = remaining_payment
                                fee_pay = diff
                            else:
                                actual_pay = remaining_payment
                        else:
                            actual_pay = remaining_payment

                    batch_summary[mgmt_no]['deposit'] += actual_pay
                    batch_summary[mgmt_no]['fee'] += fee_pay
                    
                    remaining_payment -= actual_pay

            # 4. Payment 시트에 이력 기록
            new_payment_records = []
            
            for mgmt_no, summary in batch_summary.items():
                if summary['deposit'] > 0:
                    record = {
                        "일시": now_str,
                        "관리번호": mgmt_no,
                        "구분": "입금",
                        "입금액": summary['deposit'],
                        "통화": summary['currency'],
                        "작업자": current_user,
                        "비고": f"일괄 입금 ({payment_date})"
                    }
                    if "외화입금증빙경로" in file_paths:
                        record["외화입금증빙경로"] = file_paths["외화입금증빙경로"]
                    if "송금상세경로" in file_paths:
                        record["송금상세경로"] = file_paths["송금상세경로"]
                        
                    new_payment_records.append(record)
                
                if summary['fee'] > 0:
                    new_payment_records.append({
                        "일시": now_str,
                        "관리번호": mgmt_no,
                        "구분": "수수료/조정",
                        "입금액": summary['fee'],
                        "통화": summary['currency'],
                        "작업자": current_user,
                        "비고": "잔액 탕감 처리"
                    })

            if new_payment_records:
                payment_df_new = pd.DataFrame(new_payment_records)
                dfs["payment"] = pd.concat([dfs["payment"], payment_df_new], ignore_index=True)

            # 5. 최종 재계산
            for mgmt_no in mgmt_nos:
                self.recalc_payment_status(dfs, mgmt_no)

            mgmt_str = mgmt_nos[0]
            if len(mgmt_nos) > 1: mgmt_str += f" 외 {len(mgmt_nos)-1}건"
            
            file_log = ""
            if file_paths.get("외화입금증빙경로"): file_log += " / 외화증빙"
            if file_paths.get("송금상세경로"): file_log += " / 송금상세"
            
            log_msg = f"번호 [{mgmt_str}] / 입금액 [{payment_amount:,.0f}] 처리{file_log} (재계산 완료)"
            self._add_log_to_dfs(dfs, "수금 처리", log_msg)

            return True, ""
            
        return self._execute_transaction(update)