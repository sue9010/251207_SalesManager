import os
import getpass
import pandas as pd
import openpyxl
from datetime import datetime

class DeliveryHandler:
    def __init__(self, data_manager):
        self.dm = data_manager

    def _generate_sequential_id(self, df, id_col, prefix):
        today = datetime.now().strftime("%y%m%d")
        prefix_date = f"{prefix}-{today}-"
        
        max_seq = 0
        if not df.empty and id_col in df.columns:
            def extract_seq(val):
                s = str(val)
                if s.startswith(prefix_date):
                    try: return int(s.split("-")[-1])
                    except: return 0
                return 0
            
            seqs = df[id_col].apply(extract_seq)
            if not seqs.empty: max_seq = seqs.max()
        
        next_seq = max_seq + 1
        return f"{prefix_date}{next_seq:03d}"

    def get_next_delivery_id(self):
        return self._generate_sequential_id(self.dm.df_delivery, "출고번호", "CX")

    def add_delivery(self, delivery_data: dict) -> tuple[bool, str]:
        def update(dfs):
            new_df = pd.DataFrame([delivery_data])
            if not new_df.dropna(how='all').empty:
                dfs["delivery"] = pd.concat([dfs["delivery"], new_df], ignore_index=True)
            self.dm.log_handler.add_log_to_dfs(dfs, "출고 등록", f"출고번호: {delivery_data.get('출고번호')}")
            return True, ""
        return self.dm.execute_transaction(update)

    def process_delivery(self, delivery_no, delivery_date, invoice_no, shipping_method, waybill_path, update_requests):
        def update(dfs):
            processed_items = []
            new_delivery_records = []
            
            try: current_user = getpass.getuser()
            except: current_user = "Unknown"
            
            for req in update_requests:
                idx = req["idx"]
                if idx not in dfs["data"].index: continue
                
                row_data = dfs["data"].loc[idx]
                db_qty = float(str(row_data["수량"]).replace(",", "") or 0)
                deliver_qty = min(req["deliver_qty"], db_qty)
                
                new_delivery_records.append({
                    "일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "출고번호": delivery_no, "출고일": delivery_date,
                    "관리번호": row_data.get("관리번호", ""), "품목명": row_data.get("품목명", ""),
                    "시리얼번호": req["serial_no"], "출고수량": deliver_qty,
                    "송장번호": invoice_no, "운송방법": shipping_method,
                    "작업자": current_user, "비고": "일괄 납품 처리",
                    "운송장경로": waybill_path
                })

                # 데이터 업데이트 (완전 출고 vs 부분 출고)
                is_full = abs(deliver_qty - db_qty) < 0.000001
                new_status = "완료" if row_data.get("Status") == "납품대기/입금완료" else "납품완료/입금대기"
                
                price = float(str(row_data.get("단가", 0)).replace(",", "") or 0)
                tax_rate = float(str(row_data.get("세율(%)", 0)).replace(",", "") or 0) / 100

                if is_full:
                    dfs["data"].at[idx, "Status"] = new_status
                    dfs["data"].at[idx, "출고일"] = delivery_date
                    dfs["data"].at[idx, "송장번호"] = invoice_no
                    dfs["data"].at[idx, "운송방법"] = shipping_method
                    dfs["data"].at[idx, "미수금액"] = float(str(row_data.get("합계금액", 0)).replace(",", ""))
                else:
                    remain_qty = db_qty - deliver_qty
                    supply = remain_qty * price
                    tax = supply * tax_rate
                    dfs["data"].at[idx, "수량"] = remain_qty
                    dfs["data"].at[idx, "공급가액"] = supply
                    dfs["data"].at[idx, "세액"] = tax
                    dfs["data"].at[idx, "합계금액"] = supply + tax
                    dfs["data"].at[idx, "미수금액"] = supply + tax
                    
                    new_supply = deliver_qty * price
                    new_tax = new_supply * tax_rate
                    new_row = row_data.copy()
                    new_row.update({
                        "수량": deliver_qty, "공급가액": new_supply, "세액": new_tax, "합계금액": new_supply + new_tax,
                        "미수금액": new_supply + new_tax, "Status": new_status, "출고일": delivery_date,
                        "송장번호": invoice_no, "운송방법": shipping_method,
                    })
                    if "운송장경로" in new_row: new_row["운송장경로"] = ""
                    if not pd.DataFrame([new_row]).dropna(how='all').empty:
                        dfs["data"] = pd.concat([dfs["data"], pd.DataFrame([new_row])], ignore_index=True)
                
                processed_items.append(f"{row_data.get('품목명','')} ({deliver_qty}개)")

            if new_delivery_records:
                new_df = pd.DataFrame(new_delivery_records)
                if not new_df.dropna(how='all').empty:
                    dfs["delivery"] = pd.concat([dfs["delivery"], new_df], ignore_index=True)

            log_msg = f"번호 [{row_data.get('관리번호', '')}...] 납품 처리(출고번호: {delivery_no}) / {', '.join(processed_items)}"
            self.dm.log_handler.add_log_to_dfs(dfs, "납품 처리", log_msg)
            return True, ""
            
        return self.dm.execute_transaction(update)

    def export_to_production_request(self, rows_data):
        prod_path = self.dm.production_request_path
        if not os.path.exists(prod_path):
            return False, f"생산 요청 파일을 찾을 수 없습니다.\n경로: {prod_path}"

        try:
            wb = openpyxl.load_workbook(prod_path)
            if "Data" not in wb.sheetnames:
                return False, "'Data' 시트가 존재하지 않습니다."
            ws = wb["Data"]

            added_count = 0
            updated_count = 0

            for row_data in rows_data:
                client_name = row_data.get("업체명", "")
                client_note = "-"
                if not self.dm.df_clients.empty:
                    c_row = self.dm.df_clients[self.dm.df_clients["업체명"] == client_name]
                    if not c_row.empty:
                        val = c_row.iloc[0].get("특이사항", "-")
                        if str(val) != "nan" and val: client_note = str(val)

                mgmt_no = str(row_data.get("관리번호", ""))
                model_name = str(row_data.get("모델명", ""))
                desc = str(row_data.get("Description", ""))
                order_date = row_data.get("수주일", "-")
                if not order_date or order_date == "nan": order_date = "-"

                mapping_values = [
                    mgmt_no,                    # A
                    client_name,                # B
                    model_name,                 # C
                    desc,                       # D
                    row_data.get("수량", 0),    # E
                    row_data.get("주문요청사항", "-"), # F
                    client_note,                # G
                    order_date,                 # H
                    "-",                        # I
                    "-",                        # J
                    "-",                        # K
                    "-",                        # L
                    "-",                        # M
                    "생산 접수",                # N
                    "-",                        # O
                    "-"                         # P
                ]

                target_row_idx = None
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    curr_mgmt = str(row[0]) if row[0] else ""
                    curr_model = str(row[2]) if row[2] else ""
                    curr_desc = str(row[3]) if row[3] else ""
                    
                    if curr_mgmt == mgmt_no and curr_model == model_name and curr_desc == desc:
                        target_row_idx = i
                        break
                
                if target_row_idx:
                    for col_idx, val in enumerate(mapping_values, start=1):
                        ws.cell(row=target_row_idx, column=col_idx, value=val)
                    updated_count += 1
                else:
                    ws.append(mapping_values)
                    added_count += 1

            wb.save(prod_path)
            wb.close()
            return True, f"신규: {added_count}건, 업데이트: {updated_count}건"

        except PermissionError:
            return False, "생산 요청 파일이 열려있습니다. 파일을 닫고 다시 시도해주세요."
        except Exception as e:
            return False, f"생산 요청 내보내기 실패: {e}"

    def sync_production_dates(self):
        if not os.path.exists(self.dm.production_request_path):
            return

        try:
            wb = openpyxl.load_workbook(self.dm.production_request_path, data_only=True)
            if "Data" not in wb.sheetnames: return
            ws = wb["Data"]
            
            date_map = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                mgmt_no = str(row[0]) if row[0] else None
                delivery_date = row[8]
                
                if mgmt_no and delivery_date:
                    if isinstance(delivery_date, datetime):
                        date_str = delivery_date.strftime("%Y-%m-%d")
                    else:
                        date_str = str(delivery_date).strip()
                        if date_str.lower() == "nan" or date_str == "-" or not date_str:
                            continue
                            
                    date_map[mgmt_no] = date_str
            
            wb.close()
            
            if not self.dm.df_data.empty:
                if '출고예정일' not in self.dm.df_data.columns:
                    self.dm.df_data['출고예정일'] = "-"
                
                for mgmt_no, new_date in date_map.items():
                    mask = self.dm.df_data['관리번호'] == mgmt_no
                    if mask.any():
                        self.dm.df_data.loc[mask, '출고예정일'] = new_date
                        
        except Exception as e:
            print(f"생산 요청일 동기화 실패: {e}")

    def get_production_status_map(self):
        if not os.path.exists(self.dm.production_request_path):
            return {}

        try:
            wb = openpyxl.load_workbook(self.dm.production_request_path, data_only=True, read_only=True)
            if "Data" not in wb.sheetnames:
                return {}
            
            ws = wb["Data"]
            status_map = {}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 14: continue
                
                mgmt_no = str(row[0]).strip() if row[0] else None
                prod_status = str(row[13]).strip() if row[13] else "-"
                
                if mgmt_no:
                    status_map[mgmt_no] = prod_status
            
            wb.close()
            return status_map
        except Exception as e:
            print(f"생산 상태 로드 실패: {e}")
            return {}

    def get_serial_number_map(self):
        if not os.path.exists(self.dm.production_request_path):
            return {}

        try:
            wb = openpyxl.load_workbook(self.dm.production_request_path, data_only=True, read_only=True)
            if "Data" not in wb.sheetnames:
                return {}
            
            ws = wb["Data"]
            serial_map = {}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 11: continue
                
                mgmt_no = str(row[0]).strip() if row[0] else ""
                model = str(row[2]).strip() if row[2] else ""
                desc = str(row[3]).strip() if row[3] else ""
                serial = str(row[10]).strip() if row[10] else "-"
                
                key = (mgmt_no, model, desc)
                if mgmt_no: 
                    serial_map[key] = serial
            
            wb.close()
            return serial_map
        except Exception as e:
            print(f"시리얼 번호 로드 실패: {e}")
            return {}
